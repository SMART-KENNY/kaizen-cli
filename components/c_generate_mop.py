from openpyxl import load_workbook

def replace_placeholders_in_excel(file_path, replacements, output_path=None):
    """
    Replace placeholder text in an Excel file (.xlsx).
    
    Args:
        file_path (str): Path to the input Excel file.
        replacements (dict): Dictionary of placeholder → replacement text.
        output_path (str, optional): Path to save the modified file. 
                                     If None, overwrites the input file.
    """
    wb = load_workbook(file_path)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):  # only replace in text cells
                    for target, replacement in replacements.items():
                        if target in cell.value:
                            cell.value = cell.value.replace(target, replacement)

    save_path = output_path if output_path else file_path
    wb.save(save_path)
    return save_path
